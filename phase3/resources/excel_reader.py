import openpyxl
wb = openpyxl.load_workbook('SLS+data.xlsx')
sheet = wb.get_sheet_by_name('Affiliated and Related Projects')
first_names = []
last_names = []
proj_names = []
descriptions = []
designation = []
estimated = []
emails = []
for i in range(20):
    first_names.append(sheet.cell(row=i + 3, column=1).value)
    last_names.append(sheet.cell(row=i + 3, column=2).value)
    proj_names.append(sheet.cell(row=i + 3, column=3).value)
    descriptions.append(sheet.cell(row=i + 3, column=4).value)
    designation.append(sheet.cell(row=i + 3, column=5).value)
    estimated.append(sheet.cell(row=i + 3, column=6).value)
    emails.append(sheet.cell(row=i + 3, column=7).value)
f = open('data_import_4.sql', 'w')
f.write('-- PROJECT\n')
for i in range(len(first_names)):
    statement = "INSERT INTO project (name, estNum, description, advfName, advlName, advEmail, desigName) VALUES (\'{}\', {}, \'{}\', \'{}\', \'{}\', \'{}\', \'{}\');\n".format(proj_names[i], estimated[i], descriptions[i], first_names[i], last_names[i], emails[i], designation[i])
    f.write(statement)

f.write('\n-- PROJECT CATEGORY\n')
for i in range(len(proj_names)):
    statement = "INSERT INTO project_category (projectName, categoryName) VALUES (\'{}\', \'{}\');\n".format(proj_names[i], 'temp')
    f.write(statement)
f.close()